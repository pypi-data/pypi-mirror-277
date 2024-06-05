from openpyxl import load_workbook

def get_excel_data(file_path: str, sheet_name: str = None):
    """
    Get the data from excel file.
    @param file_path: str, the path of excel file.
    @param sheet_name: str, the name of sheet.
    @return: list, the data of excel file.
    """

    wb = load_workbook(file_path)
    
    sheet = wb.active

    if sheet_name is not None:
        sheet = wb[sheet_name]

    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    return data

def excel2mdtablestr(excel_file_path: str):
    """
    Convert excel file to markdown String with table format.
    @param excel_file_path: str, the path of excel file.
    @return: str, the markdown string with table format.
    """
    
    data = get_excel_data(excel_file_path)

    if data is None or len(data) == 0:
        return None
    
    # get the 1st row as header
    header = data[0]
    header_str = "|" + "|".join([str(h) for h in header]) + "|" + "\n"

    # get the other rows as content
    content = data[1:]
    for i in range(len(content)):
        content[i] = "|" + "|".join([str(c) for c in content[i]]) + "|"

    content_str = "\n".join(content)

    return header_str + "|".join(["----" for g in header]) + "\n" + content_str

def excel2jsonlist(excel_file_path: str):
    """
    Convert excel file to json list.
    The first row of excel file is the keys of json.
    The other rows of excel file are the values of json.
    Remind: the data in the ecxel file should be paste as value first in order to clear the formula! 
    @param excel_file_path: str, the path of excel file.
    @return: dict, the json data.
    """
    data = get_excel_data(excel_file_path)

    if data is None or len(data) == 0:
        return None
    
    keys = data[0]
    values = data[1:]

    json_data = []

    for v in values:
        json_data.append(dict(zip(keys, v)))

    return json_data

def jsonlist2excel(json_list: list, excel_file_path: str):
    """
    Convert json list to excel file.
    The keys of json are the 1st row of excel file.
    @param json_list: list, the json list.
    @param excel_file_path: str, the save path of excel file.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active

    # The keys of json are the 1st row of excel file.
    keys = list(json_list[0].keys())
    sheet.append(keys)

    # The values of json are the other rows of excel file.
    for json in json_list:
        values = [json[k] if k in json else None for k in keys]
        sheet.append(values)

    wb.save(excel_file_path)

excel2mardowntablestr = excel2mdtablestr
excel2json = excel2jsonlist
json2excel = jsonlist2excel

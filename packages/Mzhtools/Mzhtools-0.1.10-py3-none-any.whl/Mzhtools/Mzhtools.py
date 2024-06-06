import base64
import hashlib
import shutil
from datetime import datetime
from subprocess import PIPE, Popen
import smtplib
from dateutil import parser
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.header import Header
import requests
import pandas as pd
import xlrd
import openpyxl
import os
r"""
cd pip_Mzhtools & del /q dist && del /q Mzhtools.egg-info && python setup.py bdist_wheel && twine upload dist/*
"""


def dabao(path):
    os.makedirs(r'c:\pybuild', exist_ok=True)
    try:
        cmd_gbk(rf'rd /S /Q "c:\pybuild"&&mkdir "c:\pybuild"&&pyinstaller {path} --workpath c:\pybuild  --distpath c:\pybuild\dist')
        return
    except:
        pass
    cmd_utf(rf'rd /S /Q "c:\pybuild"&&mkdir "c:\pybuild"&&pyinstaller {path} --workpath c:\pybuild  --distpath c:\pybuild\dist')


def qywx_sendtxt(text, key, all=None):
    """
    向指定企业微信群发送文本信息
    :param text: 发送的文本内容
    :param key: 群机器人key
    :param all: 默认None，不@全体人员。all的值为真时，@全体人员
    :return: 空
    """
    if all:
        headers = {"Content-Type": "text/plain"}
        send_url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
        send_data = {
            "msgtype": "text",  # 消息类型
            "text": {
                "content": text,  # 文本内容，最长不超过2048个字节，必须是utf8编码
                "mentioned_list": ["@all"]
            }
        }
        requests.post(url=send_url, headers=headers, json=send_data)
    else:
        headers = {"Content-Type": "text/plain"}
        send_url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
        send_data = {
            "msgtype": "text",  # 消息类型
            "text": {
                "content": text,  # 文本内容，最长不超过2048个字节，必须是utf8编码
            }
        }
        requests.post(url=send_url, headers=headers, json=send_data)


def qywx_sendImg(imgPath, key):
    """
    向指定企业微信群发送图片
    :param imgPath: 图片路径
    :param key: 群机器人key
    :return: 空
    """
    url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
    with open(imgPath, "rb") as f:
        fd = f.read()
        base64Content = str(base64.b64encode(fd), "utf-8")
    with open(imgPath, "rb") as f:
        fd = f.read()
        md = hashlib.md5()
        md.update(fd)
        md5Content = md.hexdigest()
    headers = {"content-type": "application/json"}
    msg = {"msgtype": "image", "image": {"base64": base64Content, "md5": md5Content}}
    requests.post(url, headers=headers, json=msg)


def qywx_sendfile(file, key):
    """
    向指定企业微信群发送文件
    :param file: 文件路径
    :param key: 群机器人key
    :return: 空
    """
    # 获取media_id
    id_url = f'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={key}&type=file'
    files = {'file': open(file, 'rb')}
    res = requests.post(url=id_url, files=files)
    media_id = res.json()['media_id']
    # 发送文件
    webhook = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
    header = {
        "Content-Type": "application/json",
        "Charset": "UTF-8"
    }
    data = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    }
    requests.post(url=webhook, json=data, headers=header)


def qywx_sendmarkdown(title, name, data, key):
    """
    向指定企业微信群发送markdown
    :param title: markdown标题
    :param name: markdown的key值
    :param data: markdown的value值
    :param key: 群机器人key
    :return: 空
    """
    headers = {"Content-Type": "text/plain"}
    send_url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
    content = ''
    if title:
        content = title + '\n'
    for i in zip(name, data):
        content += f"> {i[0]}：<font color=\"info\">{i[1]}</font> \n"
    send_data = {
        "msgtype": "markdown",  # 消息类型，此时固定为markdown
        "markdown": {
            "content": content
        }
    }
    requests.post(url=send_url, headers=headers, json=send_data)


def delete_file(target_dir, days):
    """
    删除指定文件夹里超过指定天数的所有东西
    :param target_dir: 文件夹路径
    :param days: 超过的天数
    :return: 空
    """
    now_time = datetime.now().strftime('%Y-%m-%d')
    for (dirpath, dirnames, filenames) in os.walk(target_dir):
        for i in filenames:
            modify_time = datetime.fromtimestamp(os.path.getmtime(os.path.join(dirpath, i))).strftime('%Y-%m-%d')
            Days = int((parser.parse(now_time) - parser.parse(modify_time)).days)
            if Days > days:
                os.remove(os.path.join(dirpath, i))
        for dirname in dirnames:
            if not os.listdir(os.path.join(dirpath, dirname)):
                shutil.rmtree(os.path.join(dirpath, dirname), ignore_errors=True)


def cmd_utf(order):
    """
    utf-8编码
    :param order: 输入对应命令
    :return: outinfo, errinfo
    """
    proc = Popen(order, stdin=None, stdout=PIPE, stderr=PIPE, shell=True)
    outinfo, errinfo = proc.communicate()
    outinfo = outinfo.decode('utf-8')
    errinfo = errinfo.decode('utf-8')
    print('outinfo', outinfo)
    print("errinfo", errinfo)
    return outinfo, errinfo


def cmd_gbk(order):
    """
    gbk编码
    :param order: 输入对应命令
    :return: outinfo, errinfo
    """
    proc = Popen(order, stdin=None, stdout=PIPE, stderr=PIPE, shell=True)
    outinfo, errinfo = proc.communicate()
    outinfo = outinfo.decode('gbk')
    errinfo = errinfo.decode('gbk')
    print('outinfo', outinfo)
    print("errinfo", errinfo)
    return outinfo, errinfo


def py_run(path):
    """
    传入py文件绝对路径，在本代码调用其他代码
    :param path:
    :return: 空
    """
    try:
        print(cmd_gbk(rf"cd /d {os.path.dirname(path)} && python {os.path.basename(path)}")[0])
        return
    except:
        pass
    print(cmd_utf(rf"cd /d {os.path.dirname(path)} && python {os.path.basename(path)}")[0])


def pushplus(title, content, token="cf8736250267472e954737c221b33d23"):
    """
    推送到微信推送加
    :param title:标题
    :param content: 内容
    :param token: 没有默认自己的微信
    :return: 空
    """
    url = fr"https://www.pushplus.plus/send?token={token}&title={title}&content={content}&template=html"
    payload = {}
    files = {}
    headers = {
        'User-Agent': 'Apifox/1.0.0 (https://apifox.com)'
    }
    requests.request("GET", url, headers=headers, data=payload, files=files)


def send_email(name_list, fpath='', title='', text=''):
    """
    :param name_list: ['341994822@qq.com']，可以添加多个
    :param fpath: 传入字符串，以空格连接，邮箱附件路径，可以添加多个，默认空
    :param title: 邮件标题，没有此参数时，默认空
    :param text: 邮件正文，没有此参数时，默认空
    :return: 空
    """
    # 生成连结对象,参数分别是邮件服务器和端口号
    con = smtplib.SMTP_SSL('smtp.qq.com', 465)
    # 使用用户名和密码登录,这里密码以星号隐藏了
    con.login('191891173@qq.com', 'tvthacuvztgicaib')
    # 生成一个邮件对象，由于邮件包含文本、图片、HTML、附件等内容，
    # 所以这里用MIMEMultipart()生成邮件对象，以支持多种数据格式
    mail_obj = MIMEMultipart()
    # 生成邮件表头的内容
    mail_header = Header(title, 'utf-8').encode()
    # 主题
    mail_obj['Subject'] = mail_header
    # 发送者邮箱
    mail_obj['From'] = '191891173@qq.com <191891173@qq.com>'
    # 接收者邮箱
    mail_obj['To'] = '我'
    # 添加邮件正文
    mail_text = MIMEText(text, 'plain', 'utf-8')
    mail_obj.attach(mail_text)
    for path in fpath.strip().split(' '):
        if path.endswith('.txt'):
            # 添加txt附件
            with open(path, 'rb') as f:
                txt = f.read()
                txt = MIMEText(txt, 'base64', 'utf-8')
                txt["Content-Type"] = 'application/octet-stream'
                txt["Content-Disposition"] = 'attachment; filename="I.txt"'
                mail_obj.attach(txt)
        if path.endswith('.xlsx'):
            # 添加Excel附件
            with open(path, 'rb') as f:
                Excel = f.read()
                Excel = MIMEText(Excel, 'base64', 'utf-8')
                Excel["Content-Type"] = 'application/octet-stream'
                Excel["Content-Disposition"] = 'attachment; filename="ove.xlsx"'
                mail_obj.attach(Excel)
        if path.endswith('.zip'):
            # 添加Zip附件
            with open(path, 'rb') as f:
                Zip = f.read()
                Zip = MIMEText(Zip, 'base64', 'utf-8')
                Zip["Content-Type"] = 'application/octet-stream'
                Zip["Content-Disposition"] = 'attachment; filename="class.rar"'
                mail_obj.attach(Zip)
        if path.endswith('.png'):
            # 添加图片附件
            with open(path, 'rb') as f:
                img2 = f.read()
                img_2 = MIMEImage(img2)
                # 指定图片类型与文件名，以下语句设置图片文件以附件形式加到邮件中
                img_2['Content-Disposition'] = 'attachment;filename="flower.png"'
                # 加入到邮件中
                mail_obj.attach(img_2)
        if path.endswith('.docx'):
            # 添加word附件
            with open(path, 'rb') as f:
                doc = f.read()
                # 以数据流的形式读入文件
                doc = MIMEText(doc, 'base64', 'utf-8')
                # 以下语句设置文件以附件形式加到邮件中
                doc['Content-Disposition'] = 'attachment;filename="test.docx"'
                # 加入到邮件中
                mail_obj.attach(doc)

    # 发送邮件
    con.sendmail('191891173@qq.com', name_list, mail_obj.as_string())
    # 断开连结
    con.quit()
    print('发送邮件成功...')


def readrow_excel(path, sheet_name):
    """
    xlrd横向读取表格
    :param path: 文件路径
    :param sheet_name: 表单名
    :return: 横向表格数据
    """
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_name(sheet_name)
    data_list = [sheet.row_values(rowx=i) for i in range(sheet.nrows)]
    return data_list


def readcol_excel(path, sheet_name):
    """
    xlrd纵向读取表格
    :param path: 文件路径
    :param sheet_name: 表单名
    :return: 横向表格数据
    """
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_name(sheet_name)
    data_list = [sheet.col_values(rowx=i) for i in range(sheet.nrows)]
    return data_list


def writelist_toExcel(ls, savename, loadname=None, sheet_name=None):
    """
    按行写入表格
    :param ls: 数据列表
    :param savename: 保存的文件名
    :param loadname: 是否写入已存在表格，默认为None，即写入新表格
    :param sheet_name: 表单名，内容写进哪个表单
    :return:
    """
    if loadname == None:
        book = openpyxl.Workbook()
        sh = book.active
        for row in ls:
            sh.append(row)
        book.save(savename)
    else:
        wb = openpyxl.load_workbook(loadname)
        sheet = wb[sheet_name]
        for row in ls:
            sheet.append(row)
        wb.save(savename)


def pd_read_excel(path, sheet_name):
    datalist = pd.read_excel(path, sheet_name, header=None, index_col=None)
    LS = []
    for j in range(len(datalist)):
        ls = []
        for i in datalist.columns:
            ls.append(datalist.iloc[j][i])
        LS.append(ls)
    return LS


def pd_writelist_toExcel(ls, savename, sheet_name):
    if not os.path.exists(savename):
        book = openpyxl.Workbook()
        sh = book.active
        sh.title = sheet_name
        book.save(savename)
    df = pd.DataFrame(ls)
    with pd.ExcelWriter(savename, datetime_format="YYYY-MM-DD") as writer:
        df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)


def pd_writelist_appendExcel(ls, savename, sheet_name):
    data = pd_read_excel(savename, sheet_name)
    ls = data + ls
    df = pd.DataFrame(ls)
    with pd.ExcelWriter(savename, datetime_format="YYYY-MM-DD") as writer:
        df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
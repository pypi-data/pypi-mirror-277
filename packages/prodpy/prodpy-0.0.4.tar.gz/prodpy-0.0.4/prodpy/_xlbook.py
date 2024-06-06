import contextlib

from difflib import SequenceMatcher

import numpy
import openpyxl

from .datum.frame._batch import Batch

from .directory._browser import Browser

class XlBook(Browser):

    def __init__(self,write=True,**kwargs):

        super().__init__(**kwargs)

        if write:
            self.book = openpyxl.Workbook()

        self.sheets = {}

    def split(self,sheetname,intFlag=False):
        """It should split to frames based on all None lists.
        
        intFlag     : boolean to set integers as unwanted element.
                     - If row contains only integers, it will be removed.
                     - None is unwanted element by default.
                     - If row contains nones and integers, it will be removed too.

        """

        frames = []
        
        subrows,prevRowNoneFlag = [],True

        for row in self.sheets[sheetname]:

            unwanted_element_count = countnone(row)

            if intFlag:
                unwanted_element_count += countint(row)

            if unwanted_element_count==len(row):
                if not prevRowNoneFlag:
                    frames.append(Batch(subrows))
                subrows,prevRowNoneFlag = [],True

            else:
                subrows.append(row)
                prevRowNoneFlag = False

        return frames

    def addsheets(self,*args):

        args = list(args)

        try:
            sheet = self.book["Sheet"]
        except KeyError:
            pass
        else:
            sheet.title = args.pop(0)
        finally:
            [self.book.create_sheet(title) for title in args]

    def addrows(self,sheetname,rows):

        sheet = self.book[sheetname]

        for row in rows:
            sheet.append(row)

    def save(self):

        self.book.save(self.filepath)

def loadxl(*args,sheetnames=None,command=False,**kwargs):
    """
    Returns an instance of textio.XlBook. If a filepath is specified, the instance
    represents the file.
    
    Arguments:
        filepath {str} -- path to the given excel file

    Keyword Arguments:
        homedir {str} -- path to the home (output) directory
        filedir {str} -- path to the file (input) directory
    
    Returns:
        textio.XlBook -- an instance of textio.XlBook filled with excel file text.

    """

    if len(args)==1:
        filepath = args[0]
    elif len(args)>1:
        raise "The function does not take more than one positional argument."

    # It creates an empty textio.XlBook instance.
    nullbook = XlBook(filepath=filepath,write=False,**kwargs)

    # It reads excel file and returns textio.XlBook instance.
    fullbook = XlWorm(nullbook,sheetnames=sheetnames).xlbook

    if command:
        print("Loaded {}".format(fullbook.filepath))

    return fullbook

class XlWorm():

    def __init__(self,xlbook,sheetnames=None,**kwargs):

        self.xlbook = xlbook

        search = pop(kwargs,"search",False)

        min_row = pop(kwargs,"min_row",1)
        min_col = pop(kwargs,"min_col",1)

        max_row = pop(kwargs,"max_row",None)
        max_col = pop(kwargs,"max_col",None)

        with XlWorm.xlopen(self.xlbook.filepath) as self.book:

            if sheetnames is None:
                sheetnames = self.book.sheetnames
            else:
                sheetnames = self.get_sheetname(name=sheetnames,search=search)

            for sheetname in sheetnames:

                rows = self.load(sheetname,min_row,min_col,max_row,max_col)

                self.xlbook.sheets[sheetname] = rows

    def get_sheetname(self,name=None,search=False):

        if isinstance(name,int):
            sheetnames = [self.book.sheetnames[name]]
        elif isinstance(name,str) and search:
            mathcscore = [SequenceMatcher(None,sheetname,name).ratio() for sheetname in self.book.sheetnames]
            sheetnames = [self.book.sheetnames[mathcscore.index(max(mathcscore))]]
        elif isinstance(name,str):
            if name not in self.book.sheetnames:
                raise ValueError(f"'{name}' could not be found in the xlbook, try search=True.")
            sheetnames = [name]
        elif hasattr(name,"len"):
            sheetnames = [self.get_sheetname(name=sheetname,search=search) for sheetname in name]
        else:
            raise TypeError(f"Expected sheetnames is either none, int or str, but the input type is {type(sheetnames)}.")

        return sheetnames

    def load(self,sheetname,min_row=1,min_col=1,max_row=None,max_col=None):
        """It reads provided excel worksheet and returns it as a frame."""

        rows = self.book[sheetname].iter_rows(
            min_row=min_row,max_row=max_row,
            min_col=min_col,max_col=max_col,
            values_only=True)

        return [list(row) for row in rows]

    @contextlib.contextmanager
    def xlopen(filepath):
        xlbook = openpyxl.load_workbook(filepath,read_only=True,data_only=True,keep_links=False)
        try:
            yield xlbook
        finally:
            xlbook._archive.close()

def pop(kwargs,key,default=None):

    try:
        return kwargs.pop(key)
    except KeyError:
        return default

def countnone(_list):

    return _list.count(None)

def countint(_list):

    counter = 0

    for value in _list:

        if isinstance(value,int):
            counter += 1

    return counter